import openpyxl as op
import tkinter as tk
import datetime
import csv
from tkinter.filedialog import askopenfilename
from tkinter.filedialog import asksaveasfilename


def date_handler(patient_date):
    if type(patient_date) == datetime.datetime:
        month = patient_date.strftime("%m")
        day = patient_date.strftime("%d")
        year = patient_date.strftime("%Y")
        return f"{month}/{day}/{year}"
    current_year = datetime.date.today().strftime("%y")
    temp = ""
    flag = False
    for c in patient_date:
        if c == "-":
            flag = False
            break
        elif c == "/":
            flag = True
            break
    if flag:
        # mm/dd/yyyy
        temp = patient_date.split("/")
    else:
        # mm-dd-yyyy
        temp = patient_date.split("-")
    month = temp[0]
    day = temp[1]
    year = temp[2]
    if len(month) == 1 and int(month) < 10:
        month = "0" + month
    if len(day) == 1 and int(day) < 10:
        day = "0" + day
    if len(year) == 2 and int(year) > int(current_year):
        year = "19" + year
    elif len(year) == 2 and int(year) <= int(current_year):
        year = "20" + year
    return f"{month}/{day}/{year}"


class SecondDoseScreener:
    """GUI that allows the user to upload an excel or CSV, and scans the file for patients who have
    currently received only one dose a two-shot series .  The user can choose to see the results in
    the text window, and/or save the scrubbed file for further use."""

    def __init__(self, parent):
        self.parent = parent
        self.parent.title("Next Dose Vaccine Screener")

        self.filename = None
        self.wb = None  # the workbook

        self.patients = set()

        self.text = tk.Text(self.parent)
        self.text.pack()
        # provide instructions
        self.instructions()

        # load file button
        self.load_button = tk.Button(self.parent, text="LOAD FILE", command=self.load)
        self.load_button.pack()

        # run script button
        self.script_button = tk.Button(self.parent, text="RUN SCRIPT", command=self.script_python)
        self.script_button.pack()

        # display file contents button
        self.display_button = tk.Button(self.parent, text="DISPLAY CONTENTS", command=self.display)
        self.display_button.pack()

        # save data button
        self.save_button = tk.Button(self.parent, text='SAVE DATA', command=self.file_save)
        self.save_button.pack()

    def instructions(self):
        # self.winfo_toplevel().title("COVID-19 Vaccine Dose 2 Generator")
        self.text.insert("end", "Step 1: Load the excel/csv file containing patient encounter data" + "\n")
        self.text.insert("end", "Step 2: Run the script to screen out patients who have received \u2265 2 (two)"
                                " shots" + "\n")
        self.text.insert("end", "Step 3: Display contents to verify file loaded and scrubbed successfully" + "\n")
        self.text.insert("end", "Step 4: Save the scrubbed data (optional)")

    def load(self):
        name = askopenfilename(filetypes=[('Excel', ('*.xls', '*.xslm', '*.xlsx')), ('CSV', '*.csv',)])
        # convert csv to xlsx file
        if name.endswith(".csv"):
            temp = op.Workbook()
            temp_wb = temp.active
            with open(name, "r") as f:
                reader = csv.reader(f)
                for row in reader:
                    temp_wb.append(row)
            name = name.replace(".csv", ".xlsx")
            temp.save(name)
        # file is xlsx, so initialize wb to be the workbook file
        self.wb = op.load_workbook(name).active
        self.filename = name
        if self.filename is not None and self.wb is not None:
            self.text.insert("end", "\n\nFile loaded successfully" + "\n")

    def display(self):
        # ask for file if not loaded yet
        if self.wb is None:
            self.load()
        # display non-duplicate patients
        if self.wb is not None and len(self.patients) != 0:
            self.text.insert("end", "\n\nFile location and name:" + "\n")
            self.text.insert("end", self.filename + "\n\n")
            self.text.insert("end", "Total Patients Requiring Follow-up: " + str(len(self.patients)) + "\n\n" )
            self.text.insert("end", "PT Name" + "\t\t\t" + "PT DOB" + "\t\t" + "Shot Date" + "\t\t" +
                             "Vaccine Name" + "\n")
            for patient in self.patients:
                v_name = str(patient[3]).split(" ")
                v_name = v_name[0]
                self.text.insert("end", str(patient[0]) + "\t\t\t" + str(patient[1]) + "\t\t" + str(patient[2]) +
                                 "\t\t" + v_name + "\n")

    def script_python(self):
        for i in range(2, self.wb.max_row + 1):
            # obtain PT name
            name = self.wb.cell(row=i, column=1)
            name = name.value
            # obtain PT birthdate
            birthdate = self.wb.cell(row=i, column=2)
            birthdate = birthdate.value
            dob = date_handler(birthdate)
            # obtain shot date
            s_date = self.wb.cell(row=i, column=3)
            s_date = s_date.value
            shot_date = date_handler(s_date)
            # obtain vaccine name/type
            v_name = self.wb.cell(row=i, column=7)
            v_name = v_name.value
            # create a patient tuple that contains the criteria above
            patient = (name, dob, shot_date, v_name)
            # check if the patient's name and DOB are already in the self.patients set and add/remove
            self.patient_handler(patient)
        self.text.insert("end", "\n\nScript run completed successfully.")

    def patient_handler(self, patient):
        patient_name = patient[0]
        patient_dob = patient[1]
        vaccine = patient[3]
        has_match = False
        for p in self.patients:
            if p[0] == patient_name and p[1] == patient_dob and p[3] == vaccine:
                has_match = True
                self.patients.remove(p)
                break
        if not has_match:
            self.patients.add(patient)

    def file_save(self):
        fname = asksaveasfilename(filetypes=(("Excel files", "*.xlsx"),
                                             ("All files", "*.*")))
        # note: this will fail unless user ends the fname with ".xlsx"
        result = op.Workbook()
        r_sheet = result.active
        # top row labels
        a1 = r_sheet.cell(row=1, column=1)
        a2 = r_sheet.cell(row=1, column=2)
        a3 = r_sheet.cell(row=1, column=3)
        a4 = r_sheet.cell(row=1, column=4)
        a1.value = "Name"
        a2.value = "DOB"
        a3.value = "Shot Date"
        a4.value = "Vaccine Name"
        # iterate through patients to be called back in patients
        start = 2
        for p in self.patients:
            # while (start <= size + 2):
            name_cell = r_sheet.cell(row=start, column=1)
            dob_cell = r_sheet.cell(row=start, column=2)
            s_date_cell = r_sheet.cell(row=start, column=3)
            v_name_cell = r_sheet.cell(row=start, column=4)
            name_cell.value = p[0]
            dob_cell.value = p[1]
            s_date_cell.value = p[2]
            v_name_cell.value = p[3]
            start += 1
        if not fname.endswith(".xlsx"):
            fname = fname + ".xlsx"
            result.save(fname)
        else:
            result.save(fname)


# --- main ---

if __name__ == '__main__':
    root = tk.Tk()
    top = SecondDoseScreener(root)
    root.mainloop()
